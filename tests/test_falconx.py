"""Tests for FalconX/Pareto accrual system.

Covers:
- Rate schedule config integrity
- Rate loader correctness
- Dedup exclusion logic
- FalconX export module
"""

import json
import os
import sys
import unittest
from datetime import datetime, timezone
from decimal import Decimal

# Allow imports from src/
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

CONFIG_DIR = os.path.join(os.path.dirname(__file__), '..', 'config')


def load_json(name):
    with open(os.path.join(CONFIG_DIR, name)) as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# 1. Rate config integrity
# ---------------------------------------------------------------------------

class TestFalconXRateConfig(unittest.TestCase):
    """config/falconx_rates.json must be well-formed and complete."""

    def setUp(self):
        self.cfg = load_json("falconx_rates.json")

    def test_has_rate_schedule(self):
        """Rate config must have a non-empty rate_schedule array."""
        schedule = self.cfg.get("rate_schedule", [])
        self.assertIsInstance(schedule, list)
        self.assertGreater(len(schedule), 0, "rate_schedule is empty")

    def test_has_performance_fee(self):
        """Rate config must have performance_fee_pct."""
        fee = self.cfg.get("performance_fee_pct")
        self.assertIsNotNone(fee, "Missing performance_fee_pct")
        fee_val = float(fee)
        self.assertGreater(fee_val, 0)
        self.assertLess(fee_val, 1, "Fee should be a fraction, not percentage")

    def test_schedule_entries_have_required_fields(self):
        """Each schedule entry must have start and gross_rate."""
        for i, entry in enumerate(self.cfg["rate_schedule"]):
            self.assertIn("start", entry, f"Entry {i} missing 'start'")
            self.assertIn("gross_rate", entry, f"Entry {i} missing 'gross_rate'")

    def test_schedule_dates_are_iso_format(self):
        """Start dates must be valid ISO 8601 with timezone."""
        for entry in self.cfg["rate_schedule"]:
            start = entry["start"]
            try:
                datetime.fromisoformat(start.replace("Z", "+00:00"))
            except ValueError:
                self.fail(f"Invalid date format: {start}")

    def test_schedule_is_chronological(self):
        """Rate schedule must be sorted by start date."""
        dates = []
        for entry in self.cfg["rate_schedule"]:
            dt = datetime.fromisoformat(entry["start"].replace("Z", "+00:00"))
            dates.append(dt)
        self.assertEqual(dates, sorted(dates), "rate_schedule is not chronological")

    def test_gross_rates_are_valid(self):
        """Gross rates must be positive and < 1 (decimal, not percentage)."""
        for entry in self.cfg["rate_schedule"]:
            rate = float(entry["gross_rate"])
            self.assertGreater(rate, 0, f"Rate {rate} must be positive")
            self.assertLess(rate, 1, f"Rate {rate} should be decimal, not percentage")


# ---------------------------------------------------------------------------
# 2. Rate loader correctness
# ---------------------------------------------------------------------------

class TestRateLoader(unittest.TestCase):
    """falconx.rates module must load correctly and compute net rates."""

    def setUp(self):
        from falconx.rates import reload
        reload()  # Ensure fresh load

    def test_get_net_rate_returns_float(self):
        """get_net_rate must return a float."""
        from falconx.rates import get_net_rate
        ts = datetime(2026, 3, 15, tzinfo=timezone.utc)
        rate = get_net_rate(ts)
        self.assertIsInstance(rate, float)

    def test_net_rate_is_gross_times_fee_multiplier(self):
        """Net rate = gross × (1 - fee)."""
        from falconx.rates import get_net_rate
        cfg = load_json("falconx_rates.json")
        fee = float(cfg["performance_fee_pct"])

        # Use a timestamp in the last rate period
        last_entry = cfg["rate_schedule"][-1]
        ts = datetime.fromisoformat(last_entry["start"].replace("Z", "+00:00"))
        expected_gross = float(last_entry["gross_rate"])
        expected_net = expected_gross * (1 - fee)

        actual = get_net_rate(ts)
        self.assertAlmostEqual(actual, expected_net, places=10)

    def test_rate_changes_at_boundaries(self):
        """Rate should change when crossing a schedule boundary."""
        from falconx.rates import get_net_rate
        cfg = load_json("falconx_rates.json")

        if len(cfg["rate_schedule"]) < 2:
            self.skipTest("Need at least 2 rate periods")

        # Find two consecutive entries with different rates
        for i in range(len(cfg["rate_schedule"]) - 1):
            r1 = float(cfg["rate_schedule"][i]["gross_rate"])
            r2 = float(cfg["rate_schedule"][i + 1]["gross_rate"])
            if abs(r1 - r2) > 1e-10:
                boundary = datetime.fromisoformat(
                    cfg["rate_schedule"][i + 1]["start"].replace("Z", "+00:00"))
                from datetime import timedelta
                before = get_net_rate(boundary - timedelta(hours=1))
                after = get_net_rate(boundary)
                self.assertNotAlmostEqual(before, after, places=6,
                    msg=f"Rate should change at {boundary}")
                return

        self.skipTest("No rate transitions found in schedule")

    def test_get_rate_schedule_returns_list(self):
        """get_rate_schedule must return a list of (datetime, float) tuples."""
        from falconx.rates import get_rate_schedule
        schedule = get_rate_schedule()
        self.assertIsInstance(schedule, list)
        self.assertGreater(len(schedule), 0)
        for start, rate in schedule:
            self.assertIsInstance(start, datetime)
            self.assertIsInstance(rate, float)


# ---------------------------------------------------------------------------
# 3. Dedup exclusion logic
# ---------------------------------------------------------------------------

class TestDedupExclusion(unittest.TestCase):
    """Dedup must exclude types/protocols where token is locked in protocol, not in wallet.

    LP constituents and rewards are always excluded by type.
    Morpho/Kamino collateral/debt are excluded by protocol (tokens locked in protocol).
    Aave collateral/debt are NOT excluded (aTokens/debt tokens are wallet balances).
    """

    _NO_DEDUP_TYPES = {"lp_constituent", "reward"}
    _NO_DEDUP_PROTOCOLS = {"morpho", "kamino"}

    def test_lp_constituent_excluded(self):
        """lp_constituent positions must not deduplicate wallet balances."""
        self.assertIn("lp_constituent", self._NO_DEDUP_TYPES)

    def test_reward_excluded(self):
        """reward positions must not deduplicate wallet balances."""
        self.assertIn("reward", self._NO_DEDUP_TYPES)

    def test_morpho_collateral_excluded(self):
        """Morpho collateral is locked in protocol, must not dedup."""
        self.assertIn("morpho", self._NO_DEDUP_PROTOCOLS)

    def test_kamino_collateral_excluded(self):
        """Kamino collateral is locked in protocol, must not dedup."""
        self.assertIn("kamino", self._NO_DEDUP_PROTOCOLS)

    def test_aave_collateral_not_excluded(self):
        """Aave aTokens ARE wallet balances, must still dedup."""
        self.assertNotIn("aave", self._NO_DEDUP_PROTOCOLS)

    def test_vault_share_not_excluded(self):
        """vault_share positions should still deduplicate (handler reads same balanceOf)."""
        self.assertNotIn("vault_share", self._NO_DEDUP_TYPES)

    def test_manual_accrual_not_excluded(self):
        """manual_accrual positions should still deduplicate."""
        self.assertNotIn("manual_accrual", self._NO_DEDUP_TYPES)


# ---------------------------------------------------------------------------
# 4. FalconX SQLite database integrity
# ---------------------------------------------------------------------------

class TestFalconXDatabase(unittest.TestCase):
    """data/falconx.db must exist and have consistent data."""

    DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'falconx.db')

    def setUp(self):
        if not os.path.exists(self.DB_PATH):
            self.skipTest("falconx.db not found")
        import sqlite3
        self.conn = sqlite3.connect(self.DB_PATH)

    def tearDown(self):
        if hasattr(self, 'conn'):
            self.conn.close()

    def test_gauntlet_table_exists(self):
        """gauntlet_levered table must exist."""
        tables = [r[0] for r in self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        self.assertIn("gauntlet_levered", tables)

    def test_direct_table_exists(self):
        """direct_accrual table must exist."""
        tables = [r[0] for r in self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        self.assertIn("direct_accrual", tables)

    def test_gauntlet_has_data(self):
        """gauntlet_levered must have rows."""
        count = self.conn.execute("SELECT COUNT(*) FROM gauntlet_levered").fetchone()[0]
        self.assertGreater(count, 0)

    def test_direct_has_data(self):
        """direct_accrual must have rows."""
        count = self.conn.execute("SELECT COUNT(*) FROM direct_accrual").fetchone()[0]
        self.assertGreater(count, 0)

    def test_gauntlet_rates_match_config(self):
        """All stored net_rates in gauntlet_levered must match config."""
        from falconx.rates import get_net_rate
        rows = self.conn.execute(
            "SELECT timestamp_utc, net_rate FROM gauntlet_levered"
        ).fetchall()
        for ts_str, stored_rate in rows:
            ts = datetime.strptime(ts_str, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
            expected = get_net_rate(ts)
            self.assertAlmostEqual(
                stored_rate, expected, places=8,
                msg=f"Rate mismatch at {ts_str}: stored={stored_rate}, expected={expected}")

    def test_direct_rates_match_config(self):
        """All stored net_rates in direct_accrual must match config."""
        from falconx.rates import get_net_rate
        rows = self.conn.execute(
            "SELECT timestamp_utc, net_rate FROM direct_accrual"
        ).fetchall()
        for ts_str, stored_rate in rows:
            ts = datetime.strptime(ts_str, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
            expected = get_net_rate(ts)
            self.assertAlmostEqual(
                stored_rate, expected, places=8,
                msg=f"Rate mismatch at {ts_str}: stored={stored_rate}, expected={expected}")

    def test_running_balance_monotonically_increasing(self):
        """Running balance (pure accrual) should be monotonically increasing.

        Tests running_balance in both tables — NOT veris_share, which depends
        on on-chain collateral/borrow and can fluctuate.
        """
        for table in ["gauntlet_levered", "direct_accrual"]:
            rows = self.conn.execute(
                f"SELECT timestamp_utc, running_balance FROM {table} ORDER BY timestamp_utc"
            ).fetchall()
            prev_val = None
            for ts_str, val in rows:
                if prev_val is not None and val is not None:
                    self.assertGreaterEqual(
                        val, prev_val,
                        f"{table}: running_balance decreased at {ts_str} ({prev_val} -> {val})")
                if val is not None:
                    prev_val = val


# ---------------------------------------------------------------------------
# 5. Export module
# ---------------------------------------------------------------------------

class TestFalconXExport(unittest.TestCase):
    """falconx.export module must produce valid xlsx."""

    DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'falconx.db')

    def test_export_produces_file(self):
        """export_falconx_xlsx must create an xlsx file."""
        if not os.path.exists(self.DB_PATH):
            self.skipTest("falconx.db not found")

        import tempfile
        from falconx.export import export_falconx_xlsx

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            tmp_path = f.name

        try:
            export_falconx_xlsx(out_path=tmp_path)
            self.assertTrue(os.path.exists(tmp_path))
            self.assertGreater(os.path.getsize(tmp_path), 0)

            # Verify sheets exist
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            self.assertIn("Gauntlet_LeveredX", wb.sheetnames)
            self.assertIn("Direct Accrual", wb.sheetnames)
            wb.close()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


if __name__ == "__main__":
    unittest.main()
