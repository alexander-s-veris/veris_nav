"""Export FalconX SQLite data to xlsx workbook.

Produces outputs/falconx_position_export.xlsx with three sheets:
Gauntlet_LeveredX, Direct Accrual, TP Changes.

Called automatically by collect.py after the FalconX updater.
"""

import os
import sqlite3

import openpyxl
from openpyxl.utils import get_column_letter

_DB_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'falconx.db')
_OUT_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'outputs', 'falconx_position_export.xlsx')

_SHEETS = [
    ("gauntlet_levered", "Gauntlet_LeveredX"),
    ("direct_accrual", "Direct Accrual"),
    ("tp_changes", "TP Changes"),
]


def export_falconx_xlsx(db_path=None, out_path=None):
    """Export all FalconX tables to an xlsx workbook.

    Args:
        db_path: Path to falconx.db. Default: data/falconx.db.
        out_path: Output xlsx path. Default: outputs/falconx_position_export.xlsx.
    """
    db_path = db_path or _DB_PATH
    out_path = out_path or _OUT_PATH

    if not os.path.exists(db_path):
        raise FileNotFoundError(f"FalconX database not found: {db_path}")

    conn = sqlite3.connect(db_path)
    wb = openpyxl.Workbook()

    for i, (table, sheet_name) in enumerate(_SHEETS):
        try:
            cursor = conn.execute(f"SELECT * FROM {table} ORDER BY timestamp_utc")
        except sqlite3.OperationalError:
            continue  # Table doesn't exist yet

        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()

        if i == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        ws.append(columns)
        for row in rows:
            ws.append(list(row))

        # Auto-width columns
        for col_idx in range(1, len(columns) + 1):
            max_len = max(
                len(str(cell.value or ''))
                for cell in ws[get_column_letter(col_idx)]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 25)

    conn.close()
    wb.save(out_path)
