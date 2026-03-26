"""
Append 5 days of March 2026 hourly data to falconx_position.xlsx.
Period: 2026-03-01 00:00 to 2026-03-05 23:00 (120 hours).
Appends to the Gauntlet_LeveredX sheet after the last existing row.
"""
import sys
import os
import time
import json
from decimal import Decimal
from datetime import datetime, timezone, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill
from web3 import Web3

with open(os.path.join(os.path.dirname(__file__), '..', '..', '.env')) as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            k, v = line.split('=', 1)
            os.environ[k.strip()] = v.strip()

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from evm import get_web3

w3 = get_web3("ethereum")

MULTICALL3 = w3.to_checksum_address("0xcA11bde05977b3631167028862bE2a173976CA11")
MORPHO = w3.to_checksum_address("0xBBBBBbbBBb9cC5e90e3b3Af64bdAF62C37EEFFCb")
GAUNTLET = w3.to_checksum_address("0x00000000d8f3d6c5DFeB2D2b5ED2276095f3aF44")
MARKET_ID = bytes.fromhex("e83d72fa5b00dcd46d9e0e860d95aa540d5ec106da5833108a9f826f21f36f52")
PARETO = w3.to_checksum_address("0x433d5b175148da32ffe1e1a37a939e1b7e79be4d")
TRANCHE_ADDR = w3.to_checksum_address("0xC26A6Fa2C37b38E549a4a1807543801Db684f99C")

VERIS_BALANCE = 2507114.7845223
VERIS_AA_TOKENS = 2473068.8259

POS_SIG = Web3.keccak(text="position(bytes32,address)")[:4]
POS_CALL = POS_SIG + MARKET_ID + bytes.fromhex(GAUNTLET[2:].lower()).rjust(32, b'\x00')
MKT_SIG = Web3.keccak(text="market(bytes32)")[:4]
MKT_CALL = MKT_SIG + MARKET_ID
SUP_SIG = Web3.keccak(text="totalSupply()")[:4]
SUP_CALL = SUP_SIG
TP_SIG = Web3.keccak(text="tranchePrice(address)")[:4]
TP_CALL = TP_SIG + bytes.fromhex(TRANCHE_ADDR[2:].lower()).rjust(32, b'\x00')

MULTICALL3_ABI = [{
    "inputs": [{"components": [
        {"name": "target", "type": "address"},
        {"name": "callData", "type": "bytes"}
    ], "name": "calls", "type": "tuple[]"}],
    "name": "aggregate",
    "outputs": [
        {"name": "blockNumber", "type": "uint256"},
        {"name": "returnData", "type": "bytes[]"}
    ],
    "stateMutability": "view", "type": "function"
}]

mc = w3.eth.contract(address=MULTICALL3, abi=MULTICALL3_ABI)

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'outputs', 'falconx_position.xlsx')

# Loan notice rates for March
# Feb: 10% (Feb 1 - Mar 3), Mar: 9.25% (Mar 3 - Apr 1)
def get_net_rate(ts):
    if ts < datetime(2026, 3, 3, tzinfo=timezone.utc):
        return 0.10 * 0.90  # Feb rate
    else:
        return 0.0925 * 0.90  # Mar rate


def find_block_near(target_ts, hint_block):
    hint_data = w3.eth.get_block(hint_block)
    hint_ts = hint_data['timestamp']
    est = hint_block + int((target_ts - hint_ts) / 12)
    latest = w3.eth.block_number
    est = max(1, min(est, latest))
    for _ in range(10):
        bd = w3.eth.get_block(est)
        if abs(bd['timestamp'] - target_ts) < 15:
            return est
        diff = target_ts - bd['timestamp']
        adj = int(diff / 12)
        if adj == 0:
            adj = 1 if diff > 0 else -1
        est += adj
        est = max(1, min(est, latest))
    return est


def query_one(block):
    calls = [
        (MORPHO, POS_CALL),
        (MORPHO, MKT_CALL),
        (GAUNTLET, SUP_CALL),
        (PARETO, TP_CALL),
    ]
    _, data = mc.functions.aggregate(calls).call(block_identifier=block)

    coll = int.from_bytes(data[0][64:96], 'big')
    borrow_shares = int.from_bytes(data[0][32:64], 'big')
    total_borrow_assets = int.from_bytes(data[1][64:96], 'big')
    total_borrow_shares = int.from_bytes(data[1][96:128], 'big')
    borrow = borrow_shares * total_borrow_assets // total_borrow_shares if total_borrow_shares > 0 else 0
    total_supply = int.from_bytes(data[2][0:32], 'big')
    tp = int.from_bytes(data[3][0:32], 'big')

    return (
        float(Decimal(str(coll)) / Decimal(10**18)),
        float(Decimal(str(borrow)) / Decimal(10**6)),
        float(Decimal(str(total_supply)) / Decimal(10**18)),
        float(Decimal(str(tp)) / Decimal(10**6)),
    )


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['Gauntlet_LeveredX']

    # Find last data row
    last_row = 1
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=1).value is not None:
            last_row = r
            break

    last_block = ws.cell(row=last_row, column=2).value
    print(f"Last existing row: {last_row}")
    print(f"Last timestamp: {ws.cell(row=last_row, column=1).value}")
    print(f"Last block: {last_block}")

    # Generate hourly timestamps: Mar 1 00:00 to Mar 5 23:00
    start = datetime(2026, 3, 1, 0, 0, 0, tzinfo=timezone.utc)
    end = datetime(2026, 3, 5, 23, 0, 0, tzinfo=timezone.utc)

    timestamps = []
    current = start
    while current <= end:
        timestamps.append(current)
        current += timedelta(hours=1)

    total = len(timestamps)
    print(f"Appending {total} hourly rows (Mar 1-5)")

    hint_block = last_block or 24567340
    t0 = time.time()

    for i, ts in enumerate(timestamps):
        target_ts = int(ts.timestamp())
        block = find_block_near(target_ts, hint_block)
        hint_block = block

        for attempt in range(5):
            try:
                coll, borrow, supply, tp = query_one(block)
                break
            except Exception as e:
                if "429" in str(e):
                    time.sleep(2 ** attempt)
                else:
                    print(f"  Error at {ts}: {str(e)[:60]}")
                    coll, borrow, supply, tp = 0, 0, 0, 0
                    break

        r = last_row + 1 + i
        net_rate = get_net_rate(ts)
        veris_pct = VERIS_BALANCE / supply if supply > 0 else 0

        # Check if AA tokens changed (collateral deposit on Mar 3 ~19:10)
        # From our data: 48,253,644 -> 55,561,262 on Mar 3
        # Veris AA tokens stay at 2,473,068.8259
        veris_aa = VERIS_AA_TOKENS

        ws.cell(row=r, column=1, value=ts.replace(tzinfo=None))  # A: Timestamp
        ws.cell(row=r, column=1).number_format = 'yyyy-mm-dd hh:mm:ss'
        ws.cell(row=r, column=2, value=block)  # B: Block
        ws.cell(row=r, column=3, value=coll)  # C: Collateral
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=4, value=borrow)  # D: Borrow
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        ws.cell(row=r, column=5, value=supply)  # E: Vault Total Supply
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=6, value=VERIS_BALANCE)  # F: Veris Balance
        ws.cell(row=r, column=6).number_format = '#,##0.00'
        ws.cell(row=r, column=7).value = f'=IF(E{r}>0,F{r}/E{r},0)'  # G: Veris %
        ws.cell(row=r, column=7).number_format = '0.0000%'
        ws.cell(row=r, column=8, value=net_rate)  # H: Net Rate
        ws.cell(row=r, column=8).number_format = '0.000%'
        ws.cell(row=r, column=9, value=veris_aa)  # I: Veris AA tokens
        ws.cell(row=r, column=9).number_format = '#,##0.0000'
        ws.cell(row=r, column=10, value=tp)  # J: Tranche Price
        # K: TP Change
        ws.cell(row=r, column=11).value = f'=J{r}<>J{r-1}'
        # L: Period (days)
        ws.cell(row=r, column=12).value = f'=A{r}-A{r-1}'
        ws.cell(row=r, column=12).number_format = '0.000000'
        # M: Opening Value = prior Running Balance
        ws.cell(row=r, column=13).value = f'=P{r-1}'
        ws.cell(row=r, column=13).number_format = '#,##0.00'
        # N: Additional deposits
        ws.cell(row=r, column=14).value = f'=IF(I{r}<>I{r-1},(I{r}-I{r-1})*J{r},0)'
        ws.cell(row=r, column=14).number_format = '#,##0.00'
        # O: Interest
        ws.cell(row=r, column=15).value = f'=M{r}*H{r}*L{r}/365'
        ws.cell(row=r, column=15).number_format = '#,##0.00'
        # P: Running Balance
        ws.cell(row=r, column=16).value = f'=M{r}+N{r}+O{r}'
        ws.cell(row=r, column=16).number_format = '#,##0.00'
        # Q: TP re-engineered
        ws.cell(row=r, column=17).value = f'=IF(I{r}>0,P{r}/I{r},0)'
        # S: Collateral USD
        ws.cell(row=r, column=19).value = f'=C{r}*J{r}'
        ws.cell(row=r, column=19).number_format = '#,##0.00'
        # T: Net
        ws.cell(row=r, column=20).value = f'=S{r}-D{r}'
        ws.cell(row=r, column=20).number_format = '#,##0.00'
        # U: Veris share
        ws.cell(row=r, column=21).value = f'=T{r}*G{r}'
        ws.cell(row=r, column=21).number_format = '#,##0.00'

        if (i + 1) % 20 == 0 or i == total - 1:
            elapsed = time.time() - t0
            rate = (i + 1) / elapsed if elapsed > 0 else 0
            eta = (total - i - 1) / rate if rate > 0 else 0
            print(f"  {i+1}/{total} | {ts.strftime('%Y-%m-%d %H:%M')} | "
                  f"coll: {coll:,.0f} | borrow: ${borrow:,.0f} | TP: {tp} | "
                  f"{rate:.1f}/s | ETA: {eta:.0f}s")

        time.sleep(0.1)

    # Save
    for _ in range(3):
        try:
            wb.save(XLSX_PATH)
            print(f"\nAppended {total} rows. Saved to {XLSX_PATH}")
            break
        except PermissionError:
            wb.save(XLSX_PATH + ".tmp.xlsx")
            print(f"\nSaved to {XLSX_PATH}.tmp.xlsx (close Excel and rename)")
            break


if __name__ == "__main__":
    main()
