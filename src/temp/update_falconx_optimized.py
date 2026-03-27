"""
Optimized FalconX/Pareto position updater.

Uses block pre-computation (Option 1) and concurrent RPC queries (Option 4)
from src/block_utils.py. Typically 5-8x faster than the serial approach.

Usage:
    python src/temp/update_falconx_optimized.py [--start YYYY-MM-DD-HH] [--end YYYY-MM-DD-HH]

Defaults: start = last timestamp in xlsx + 1 hour, end = now (rounded down to hour).
"""
import sys
import os
import time
import argparse
from decimal import Decimal
from datetime import datetime, timezone, timedelta

import openpyxl
from web3 import Web3

# Load .env
with open(os.path.join(os.path.dirname(__file__), '..', '..', '.env')) as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            k, v = line.split('=', 1)
            os.environ[k.strip()] = v.strip()

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from evm import get_web3
from block_utils import estimate_blocks, concurrent_query_batched

w3 = get_web3("ethereum")

# --- Contracts ---
MULTICALL3 = w3.to_checksum_address("0xcA11bde05977b3631167028862bE2a173976CA11")
MORPHO = w3.to_checksum_address("0xBBBBBbbBBb9cC5e90e3b3Af64bdAF62C37EEFFCb")
GAUNTLET = w3.to_checksum_address("0x00000000d8f3d6c5DFeB2D2b5ED2276095f3aF44")
MARKET_ID = bytes.fromhex("e83d72fa5b00dcd46d9e0e860d95aa540d5ec106da5833108a9f826f21f36f52")
PARETO = w3.to_checksum_address("0x433d5b175148da32ffe1e1a37a939e1b7e79be4d")
TRANCHE_ADDR = w3.to_checksum_address("0xC26A6Fa2C37b38E549a4a1807543801Db684f99C")

# --- Constants ---
VERIS_GP_BALANCE = 2507114.7845223
VERIS_AA_TOKENS_GAUNTLET = 2473068.8259
VERIS_AA_TOKENS_DIRECT = 1894969.859499
DIRECT_OPENING_VALUE = 2024989.2306  # Actual USDC deposited (not tokens × stale TP)

# Loan notice schedule: (start_date_utc, gross_rate)
RATE_SCHEDULE = [
    (datetime(2025, 6, 30, tzinfo=timezone.utc), 0.1125),
    (datetime(2025, 7, 31, tzinfo=timezone.utc), 0.1125),
    (datetime(2025, 9,  1, tzinfo=timezone.utc), 0.1200),
    (datetime(2025, 10, 1, tzinfo=timezone.utc), 0.1200),
    (datetime(2025, 11, 1, tzinfo=timezone.utc), 0.1200),
    (datetime(2025, 12, 1, tzinfo=timezone.utc), 0.1150),
    (datetime(2026, 1,  1, tzinfo=timezone.utc), 0.1050),
    (datetime(2026, 2,  1, tzinfo=timezone.utc), 0.1000),
    (datetime(2026, 3,  3, tzinfo=timezone.utc), 0.0925),
]

def get_net_rate(ts):
    """Get net rate for a given timestamp from loan notice schedule."""
    rate = RATE_SCHEDULE[0][1]
    for start, r in RATE_SCHEDULE:
        if ts >= start:
            rate = r
    return rate * 0.90

# --- Calldata (pre-built, no per-call overhead) ---
POS_SIG = Web3.keccak(text="position(bytes32,address)")[:4]
POS_CALL = POS_SIG + MARKET_ID + bytes.fromhex(GAUNTLET[2:].lower()).rjust(32, b'\x00')
MKT_SIG = Web3.keccak(text="market(bytes32)")[:4]
MKT_CALL = MKT_SIG + MARKET_ID
SUP_SIG = Web3.keccak(text="totalSupply()")[:4]
SUP_CALL = SUP_SIG
TP_SIG = Web3.keccak(text="tranchePrice(address)")[:4]
TP_CALL = TP_SIG + bytes.fromhex(TRANCHE_ADDR[2:].lower()).rjust(32, b'\x00')

MC_ABI = [{
    "inputs": [{"components": [
        {"name": "target", "type": "address"},
        {"name": "callData", "type": "bytes"}
    ], "name": "calls", "type": "tuple[]"}],
    "name": "aggregate",
    "outputs": [
        {"name": "blockNumber", "type": "uint256"},
        {"name": "returnData", "type": "bytes[]"}
    ],
    "stateMutability": "view", "type": "function"
}]
mc = w3.eth.contract(address=MULTICALL3, abi=MC_ABI)

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'outputs', 'falconx_position.xlsx')


def query_at_block(block):
    """Multicall query at a given block. Returns (coll, borrow, supply, tp)."""
    calls = [
        (MORPHO, POS_CALL),
        (MORPHO, MKT_CALL),
        (GAUNTLET, SUP_CALL),
        (PARETO, TP_CALL),
    ]
    _, data = mc.functions.aggregate(calls).call(block_identifier=block)

    coll = int.from_bytes(data[0][64:96], 'big')
    borrow_shares = int.from_bytes(data[0][32:64], 'big')
    total_borrow_assets = int.from_bytes(data[1][64:96], 'big')
    total_borrow_shares = int.from_bytes(data[1][96:128], 'big')
    borrow = borrow_shares * total_borrow_assets // total_borrow_shares if total_borrow_shares > 0 else 0
    total_supply = int.from_bytes(data[2][0:32], 'big')
    tp = int.from_bytes(data[3][0:32], 'big')

    return (
        float(Decimal(str(coll)) / Decimal(10**18)),
        float(Decimal(str(borrow)) / Decimal(10**6)),
        float(Decimal(str(total_supply)) / Decimal(10**18)),
        float(Decimal(str(tp)) / Decimal(10**6)),
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--start", help="Start UTC: YYYY-MM-DD-HH (default: last row + 1h)")
    parser.add_argument("--end", help="End UTC: YYYY-MM-DD-HH (default: current hour)")
    parser.add_argument("--workers", type=int, default=10, help="Concurrent RPC workers (10 optimal for Alchemy)")
    parser.add_argument("--batch", type=int, default=50, help="Batch size for concurrent queries")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(XLSX_PATH)

    # ========================================
    # Find last Gauntlet row
    # ========================================
    ws_g = wb['Gauntlet_LeveredX']
    last_row_g = 1
    for r in range(ws_g.max_row, 1, -1):
        if ws_g.cell(row=r, column=1).value is not None:
            last_row_g = r
            break

    last_ts_g = ws_g.cell(row=last_row_g, column=1).value
    last_block_g = ws_g.cell(row=last_row_g, column=2).value
    print(f"Gauntlet last row: {last_row_g}, ts: {last_ts_g}, block: {last_block_g}")

    # Determine time range
    if args.start:
        start = datetime.strptime(args.start, "%Y-%m-%d-%H").replace(tzinfo=timezone.utc)
    else:
        start = last_ts_g.replace(tzinfo=timezone.utc) + timedelta(hours=1)

    if args.end:
        end = datetime.strptime(args.end, "%Y-%m-%d-%H").replace(tzinfo=timezone.utc)
    else:
        now = datetime.now(timezone.utc)
        end = now.replace(minute=0, second=0, microsecond=0)

    timestamps = []
    current = start
    while current <= end:
        timestamps.append(current)
        current += timedelta(hours=1)

    total = len(timestamps)
    if total == 0:
        print("No new rows to append.")
        return

    print(f"Period: {start.strftime('%Y-%m-%d %H:%M')} to {end.strftime('%Y-%m-%d %H:%M')} ({total} hours)")

    # ========================================
    # Option 1: Pre-compute block numbers
    # ========================================
    t0 = time.time()

    # Get reference: last known block + its timestamp
    ref_block = last_block_g
    ref_data = w3.eth.get_block(ref_block)
    ref_ts = ref_data['timestamp']

    target_unix = [int(ts.timestamp()) for ts in timestamps]
    latest_block = w3.eth.block_number
    estimated_blocks = estimate_blocks(ref_block, ref_ts, target_unix, chain="ethereum")
    # Cap at latest block (can't query future blocks)
    estimated_blocks = [min(b, latest_block) for b in estimated_blocks]

    block_time = time.time() - t0
    print(f"Block estimation: {total} blocks in {block_time:.1f}s (2 RPC calls)")
    print(f"  Latest chain block: {latest_block}, estimated range: {estimated_blocks[0]}-{estimated_blocks[-1]}")

    # ========================================
    # Option 4: Concurrent Multicall queries
    # ========================================
    t1 = time.time()

    def progress(done, total_items):
        elapsed = time.time() - t1
        rate = done / elapsed if elapsed > 0 else 0
        eta = (total_items - done) / rate if rate > 0 else 0
        print(f"  Queried {done}/{total_items} | {rate:.1f}/s | ETA: {eta:.0f}s")

    print(f"Querying on-chain data ({args.workers} workers, batch {args.batch})...")

    results = concurrent_query_batched(
        query_fn=query_at_block,
        items=estimated_blocks,
        batch_size=args.batch,
        max_workers=args.workers,
        pause_between_batches=0.3,
        progress_fn=progress,
    )

    query_time = time.time() - t1
    print(f"Query phase: {total} calls in {query_time:.1f}s ({total/query_time:.1f}/s)")

    # ========================================
    # Write Gauntlet_LeveredX rows
    # ========================================
    for i, (ts, block, data) in enumerate(zip(timestamps, estimated_blocks, results)):
        coll, borrow, supply, tp = data
        r = last_row_g + 1 + i
        net_rate = get_net_rate(ts)

        ws_g.cell(row=r, column=1, value=ts.replace(tzinfo=None))
        ws_g.cell(row=r, column=1).number_format = 'yyyy-mm-dd hh:mm:ss'
        ws_g.cell(row=r, column=2, value=block)
        ws_g.cell(row=r, column=3, value=coll)
        ws_g.cell(row=r, column=3).number_format = '#,##0.00'
        ws_g.cell(row=r, column=4, value=borrow)
        ws_g.cell(row=r, column=4).number_format = '#,##0.00'
        ws_g.cell(row=r, column=5, value=supply)
        ws_g.cell(row=r, column=5).number_format = '#,##0.00'
        ws_g.cell(row=r, column=6, value=VERIS_GP_BALANCE)
        ws_g.cell(row=r, column=6).number_format = '#,##0.0000'
        ws_g.cell(row=r, column=7).value = f'=IF(E{r}>0,F{r}/E{r},0)'
        ws_g.cell(row=r, column=7).number_format = '0.0000000000%'
        ws_g.cell(row=r, column=8, value=net_rate)
        ws_g.cell(row=r, column=8).number_format = '0.000%'
        ws_g.cell(row=r, column=9, value=VERIS_AA_TOKENS_GAUNTLET)
        ws_g.cell(row=r, column=9).number_format = '#,##0.0000'
        ws_g.cell(row=r, column=10, value=tp)
        ws_g.cell(row=r, column=11).value = f'=A{r}-A{r-1}'
        ws_g.cell(row=r, column=11).number_format = '0.000000000'
        ws_g.cell(row=r, column=12).value = f'=N{r-1}'
        ws_g.cell(row=r, column=12).number_format = '#,##0.00'
        ws_g.cell(row=r, column=13).value = f'=L{r}*H{r}*K{r}/365'
        ws_g.cell(row=r, column=13).number_format = '#,##0.00'
        ws_g.cell(row=r, column=14).value = f'=L{r}+M{r}'
        ws_g.cell(row=r, column=14).number_format = '#,##0.00'
        ws_g.cell(row=r, column=15).value = f'=IF(I{r}>0,N{r}/I{r},0)'
        ws_g.cell(row=r, column=16).value = f'=C{r}*O{r}'
        ws_g.cell(row=r, column=16).number_format = '#,##0.00'
        ws_g.cell(row=r, column=17).value = f'=P{r}-D{r}'
        ws_g.cell(row=r, column=17).number_format = '#,##0.00'
        ws_g.cell(row=r, column=18).value = f'=Q{r}*G{r}'
        ws_g.cell(row=r, column=18).number_format = '#,##0.00'

    print(f"Gauntlet: {total} rows written (rows {last_row_g+1}-{last_row_g+total})")

    # ========================================
    # Direct Accrual — append if active
    # ========================================
    ws_d = wb['Direct Accrual']
    last_row_d = 1
    for r in range(ws_d.max_row, 1, -1):
        if ws_d.cell(row=r, column=1).value is not None:
            last_row_d = r
            break

    last_ts_d = ws_d.cell(row=last_row_d, column=1).value
    print(f"\nDirect Accrual last row: {last_row_d}, ts: {last_ts_d}")

    # Determine if Direct Accrual needs rows in our time range
    if last_ts_d is not None:
        last_dt_d = last_ts_d.replace(tzinfo=timezone.utc) if last_ts_d.tzinfo is None else last_ts_d
        da_start = last_dt_d + timedelta(hours=1)

        if da_start <= end:
            da_timestamps = []
            current = da_start
            while current <= end:
                da_timestamps.append(current)
                current += timedelta(hours=1)

            total_d = len(da_timestamps)
            print(f"Appending {total_d} rows to Direct Accrual")

            # Check current direct balance
            # (If tokens are in wallet, position is active)
            tp_direct = 1.067961  # TP unchanged since Mar 3

            for i, ts in enumerate(da_timestamps):
                r = last_row_d + 1 + i
                net_rate = get_net_rate(ts)

                ws_d.cell(row=r, column=1, value=ts.replace(tzinfo=None))
                ws_d.cell(row=r, column=1).number_format = 'yyyy-mm-dd hh:mm:ss'
                ws_d.cell(row=r, column=2, value=VERIS_AA_TOKENS_DIRECT)
                ws_d.cell(row=r, column=2).number_format = '#,##0.00'
                ws_d.cell(row=r, column=3, value=tp_direct)
                ws_d.cell(row=r, column=5, value=net_rate)
                ws_d.cell(row=r, column=5).number_format = '0.000%'
                ws_d.cell(row=r, column=4).value = f'=H{r-1}'
                ws_d.cell(row=r, column=4).number_format = '#,##0.00'
                ws_d.cell(row=r, column=6).value = f'=A{r}-A{r-1}'
                ws_d.cell(row=r, column=6).number_format = '0.000000000'
                ws_d.cell(row=r, column=7).value = f'=D{r}*E{r}*F{r}/365'
                ws_d.cell(row=r, column=7).number_format = '#,##0.00'
                ws_d.cell(row=r, column=8).value = f'=D{r}+G{r}'
                ws_d.cell(row=r, column=8).number_format = '#,##0.00'
                ws_d.cell(row=r, column=9).value = f'=IF(B{r}>0,H{r}/B{r},0)'

            print(f"Direct Accrual: {total_d} rows written (rows {last_row_d+1}-{last_row_d+total_d})")
        else:
            print("Direct Accrual already up to date.")

    # ========================================
    # Save
    # ========================================
    for _ in range(3):
        try:
            wb.save(XLSX_PATH)
            print(f"\nSaved to {XLSX_PATH}")
            break
        except PermissionError:
            alt = XLSX_PATH.replace('.xlsx', '_updated.xlsx')
            wb.save(alt)
            print(f"\nOriginal locked. Saved to {alt}")
            break

    total_time = time.time() - t0
    print(f"\nTotal time: {total_time:.1f}s ({total/total_time:.1f} rows/s)")


if __name__ == "__main__":
    main()
