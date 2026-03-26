"""
Collect hourly Gauntlet/Morpho data using Multicall3.
Single RPC call per hour: position + market + totalSupply.
Rate-limited with retry. Saves every 100 rows.

Output: outputs/falconx_position.xlsx
"""
import sys
import os
import time
from decimal import Decimal
from datetime import datetime, timezone, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill
from web3 import Web3

with open(os.path.join(os.path.dirname(__file__), '..', '..', '.env')) as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            k, v = line.split('=', 1)
            os.environ[k.strip()] = v.strip()

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from evm import get_web3

w3 = get_web3("ethereum")

MULTICALL3 = w3.to_checksum_address("0xcA11bde05977b3631167028862bE2a173976CA11")
MORPHO = w3.to_checksum_address("0xBBBBBbbBBb9cC5e90e3b3Af64bdAF62C37EEFFCb")
GAUNTLET = w3.to_checksum_address("0x00000000d8f3d6c5DFeB2D2b5ED2276095f3aF44")
MARKET_ID_BYTES = bytes.fromhex("e83d72fa5b00dcd46d9e0e860d95aa540d5ec106da5833108a9f826f21f36f52")
VERIS_BALANCE = 2507114.7845223

# Pre-encode calldata
POS_SIG = Web3.keccak(text="position(bytes32,address)")[:4]
POS_CALL = POS_SIG + MARKET_ID_BYTES + bytes.fromhex(GAUNTLET[2:].lower()).rjust(32, b'\x00')

MKT_SIG = Web3.keccak(text="market(bytes32)")[:4]
MKT_CALL = MKT_SIG + MARKET_ID_BYTES

SUP_SIG = Web3.keccak(text="totalSupply()")[:4]
SUP_CALL = SUP_SIG

MULTICALL3_ABI = [{
    "inputs": [{"components": [
        {"name": "target", "type": "address"},
        {"name": "callData", "type": "bytes"}
    ], "name": "calls", "type": "tuple[]"}],
    "name": "aggregate",
    "outputs": [
        {"name": "blockNumber", "type": "uint256"},
        {"name": "returnData", "type": "bytes[]"}
    ],
    "stateMutability": "view", "type": "function"
}]

mc = w3.eth.contract(address=MULTICALL3, abi=MULTICALL3_ABI)

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'outputs',
                           'falconx_position.xlsx')


def estimate_blocks(start_ts, end_ts):
    """Estimate block numbers for each hour using linear interpolation."""
    ref_block = 23283000
    ref_ts = w3.eth.get_block(ref_block)['timestamp']

    result = []
    ts = start_ts
    while ts <= end_ts:
        est = ref_block + int((ts - ref_ts) / 12)
        result.append((ts, max(est, 1)))
        ts += 3600
    return result


def query_one(block):
    """Single Multicall3 call: position + market + totalSupply."""
    calls = [
        (MORPHO, POS_CALL),
        (MORPHO, MKT_CALL),
        (GAUNTLET, SUP_CALL),
    ]
    _, data = mc.functions.aggregate(calls).call(block_identifier=block)

    # Position: supplyShares(uint256) + borrowShares(uint128 padded) + collateral(uint128 padded)
    coll = int.from_bytes(data[0][64:96], 'big')
    borrow_shares = int.from_bytes(data[0][32:64], 'big')

    # Market: 6 x uint128 (padded to 32 bytes each)
    total_borrow_assets = int.from_bytes(data[1][64:96], 'big')
    total_borrow_shares = int.from_bytes(data[1][96:128], 'big')

    borrow = borrow_shares * total_borrow_assets // total_borrow_shares if total_borrow_shares > 0 else 0

    # Total supply
    total_supply = int.from_bytes(data[2][0:32], 'big')

    return (
        float(Decimal(str(coll)) / Decimal(10**18)),
        float(Decimal(str(borrow)) / Decimal(10**6)),
        float(Decimal(str(total_supply)) / Decimal(10**18)),
    )


def save_xlsx(wb):
    for _ in range(3):
        try:
            wb.save(OUTPUT_PATH)
            return
        except PermissionError:
            wb.save(OUTPUT_PATH + ".tmp.xlsx")
            return


def main():
    start_ts = int(datetime(2025, 9, 3, 17, 0, 0, tzinfo=timezone.utc).timestamp())
    end_ts = int(datetime(2026, 2, 28, 23, 59, 59, tzinfo=timezone.utc).timestamp())

    print("Estimating blocks...")
    block_list = estimate_blocks(start_ts, end_ts)
    total = len(block_list)
    print(f"Total points: {total}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hourly Data"

    bold = Font(bold=True)
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for c, h in enumerate(["Timestamp (UTC)", "Block",
                            "Collateral (AA_FalconXUSDC)", "Borrow (USDC)",
                            "Vault Total Supply", "Veris Balance", "Veris %"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = hfill

    for c, w in enumerate([20, 12, 25, 18, 20, 18, 12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    t0 = time.time()

    for i, (ts, block) in enumerate(block_list):
        dt = datetime.fromtimestamp(ts, tz=timezone.utc)
        r = i + 2

        for attempt in range(5):
            try:
                coll, borrow, supply = query_one(block)
                break
            except Exception as e:
                if "429" in str(e) or "Too Many" in str(e):
                    time.sleep(2 ** attempt)
                else:
                    print(f"  Error at {dt.strftime('%Y-%m-%d %H:%M')} block {block}: {str(e)[:60]}")
                    coll, borrow, supply = 0, 0, 0
                    break

        ws.cell(row=r, column=1, value=dt.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=r, column=2, value=block)
        ws.cell(row=r, column=3, value=coll)
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=4, value=borrow)
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        ws.cell(row=r, column=5, value=supply)
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=6, value=VERIS_BALANCE)
        ws.cell(row=r, column=6).number_format = '#,##0.00'
        ws.cell(row=r, column=7).value = f'=IF(E{r}>0,F{r}/E{r},0)'
        ws.cell(row=r, column=7).number_format = '0.0000%'

        if (i + 1) % 100 == 0 or i == total - 1:
            elapsed = time.time() - t0
            rate = (i + 1) / elapsed
            eta = (total - i - 1) / rate if rate > 0 else 0
            veris_pct = VERIS_BALANCE / supply * 100 if supply > 0 else 0
            print(f"  {i+1}/{total} ({(i+1)/total*100:.1f}%) | "
                  f"{dt.strftime('%Y-%m-%d %H:%M')} | "
                  f"coll: {coll:,.0f} | borrow: ${borrow:,.0f} | "
                  f"supply: {supply:,.0f} | V%: {veris_pct:.2f}% | "
                  f"{rate:.1f}/s | ETA: {eta/60:.1f}min")
            save_xlsx(wb)

        time.sleep(0.1)  # Rate limit: ~10/s

    save_xlsx(wb)
    print(f"\nDone in {(time.time()-t0)/60:.1f} min. Saved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
