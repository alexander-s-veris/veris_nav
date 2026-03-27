"""
Import FalconX position data from xlsx workbook into SQLite database.

One-time migration script. Reads outputs/falconx_position.xlsx and creates
data/falconx.db with two tables: gauntlet_levered and direct_accrual.

Idempotent — uses INSERT OR REPLACE, safe to re-run.

Usage:
    python src/temp/import_falconx_xlsx_to_sqlite.py
    python src/temp/import_falconx_xlsx_to_sqlite.py --export gauntlet_levered
    python src/temp/import_falconx_xlsx_to_sqlite.py --export direct_accrual
"""
import argparse
import csv
import os
import sqlite3
import sys
from datetime import datetime, timezone
from decimal import Decimal

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.join(SCRIPT_DIR, '..', '..')
XLSX_PATH = os.path.join(PROJECT_ROOT, 'outputs', 'falconx_position.xlsx')
DB_PATH = os.path.join(PROJECT_ROOT, 'data', 'falconx.db')

# Constants from update_falconx_optimized.py
VERIS_GP_BALANCE = 2507114.7845223
VERIS_AA_TOKENS_GAUNTLET = 2473068.8259
VERIS_AA_TOKENS_DIRECT = 1894969.859499
DIRECT_OPENING_VALUE = 2024989.2306

# Rate schedule
RATE_SCHEDULE = [
    (datetime(2025, 6, 30, tzinfo=timezone.utc), 0.1125),
    (datetime(2025, 7, 31, tzinfo=timezone.utc), 0.1125),
    (datetime(2025, 9,  1, tzinfo=timezone.utc), 0.1200),
    (datetime(2025, 10, 1, tzinfo=timezone.utc), 0.1200),
    (datetime(2025, 11, 1, tzinfo=timezone.utc), 0.1200),
    (datetime(2025, 12, 1, tzinfo=timezone.utc), 0.1150),
    (datetime(2026, 1,  1, tzinfo=timezone.utc), 0.1050),
    (datetime(2026, 2,  1, tzinfo=timezone.utc), 0.1000),
    (datetime(2026, 3,  3, tzinfo=timezone.utc), 0.0925),
]


def get_net_rate(ts):
    """Get net rate for a given timestamp from loan notice schedule."""
    rate = RATE_SCHEDULE[0][1]
    for start, r in RATE_SCHEDULE:
        if ts >= start:
            rate = r
    return rate * 0.90


def _ts_to_iso(ts):
    """Convert datetime to ISO string for SQLite storage."""
    if ts is None:
        return None
    if hasattr(ts, 'strftime'):
        return ts.strftime('%Y-%m-%d %H:%M:%S')
    return str(ts)


def _ts_to_aware(ts):
    """Convert a naive datetime to UTC-aware for rate lookups."""
    if ts is None:
        return None
    if hasattr(ts, 'tzinfo') and ts.tzinfo is None:
        return ts.replace(tzinfo=timezone.utc)
    return ts


def create_tables(conn):
    """Create the SQLite tables if they don't exist."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS gauntlet_levered (
            timestamp_utc TEXT PRIMARY KEY,
            block INTEGER,
            collateral REAL,
            borrow REAL,
            vault_total_supply REAL,
            veris_balance REAL,
            veris_pct REAL,
            net_rate REAL,
            veris_aa_falconx REAL,
            tranche_price REAL,
            period_days REAL,
            opening_value REAL,
            interest REAL,
            running_balance REAL,
            tp_reengineered REAL,
            collateral_usd REAL,
            net REAL,
            veris_share REAL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS direct_accrual (
            timestamp_utc TEXT PRIMARY KEY,
            token_balance REAL,
            tranche_price REAL,
            opening_value REAL,
            net_rate REAL,
            period_days REAL,
            interest REAL,
            running_balance REAL,
            tp_reengineered REAL
        )
    """)
    conn.commit()


def import_gauntlet(conn, wb):
    """Import Gauntlet_LeveredX sheet into gauntlet_levered table."""
    ws = wb['Gauntlet_LeveredX']

    # Read all rows (skip header row 1)
    raw_rows = []
    for row in ws.iter_rows(min_row=2):
        ts = row[0].value
        if ts is None:
            continue
        raw_rows.append([cell.value for cell in row])

    if not raw_rows:
        print("Gauntlet_LeveredX: no data rows found")
        return

    print(f"Gauntlet_LeveredX: {len(raw_rows)} data rows read from xlsx")

    # Process rows — compute formula values where None
    processed = []
    prev_running_balance = None
    prev_ts = None

    for row in raw_rows:
        ts = row[0]
        if isinstance(ts, str):
            ts = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
        ts_aware = _ts_to_aware(ts)
        block = row[1]
        collateral = row[2]
        borrow = row[3]
        vault_total_supply = row[4]
        veris_balance = row[5] if row[5] is not None else VERIS_GP_BALANCE
        net_rate = row[7]
        veris_aa_falconx = row[8] if row[8] is not None else VERIS_AA_TOKENS_GAUNTLET
        tranche_price = row[9]

        # Formula columns — use cached value if available, else compute
        veris_pct = row[6]
        period_days = row[10]
        opening_value = row[11]
        interest = row[12]
        running_balance = row[13]
        tp_reengineered = row[14]
        collateral_usd = row[15]
        net = row[16]
        veris_share = row[17] if len(row) > 17 else None

        # Compute veris_pct if missing
        if veris_pct is None and vault_total_supply and vault_total_supply > 0:
            veris_pct = veris_balance / vault_total_supply

        # Compute period_days if missing
        if period_days is None and prev_ts is not None:
            delta = (ts - prev_ts).total_seconds()
            period_days = delta / 86400.0

        # Compute opening_value if missing
        if opening_value is None and prev_running_balance is not None:
            opening_value = prev_running_balance

        # Compute interest if missing
        if interest is None and opening_value is not None and net_rate is not None and period_days is not None:
            interest = opening_value * net_rate * period_days / 365.0

        # Compute running_balance if missing
        if running_balance is None:
            if opening_value is not None and interest is not None:
                running_balance = opening_value + interest
            elif prev_running_balance is None and veris_aa_falconx and tranche_price:
                # First row bootstrap
                running_balance = veris_aa_falconx * tranche_price

        # Compute tp_reengineered if missing
        if tp_reengineered is None and running_balance is not None and veris_aa_falconx and veris_aa_falconx > 0:
            tp_reengineered = running_balance / veris_aa_falconx

        # Compute collateral_usd if missing
        if collateral_usd is None and collateral is not None and tp_reengineered is not None:
            collateral_usd = collateral * tp_reengineered

        # Compute net if missing
        if net is None and collateral_usd is not None and borrow is not None:
            net = collateral_usd - borrow

        # Compute veris_share if missing
        if veris_share is None and net is not None and veris_pct is not None:
            veris_share = net * veris_pct

        processed.append((
            _ts_to_iso(ts), block, collateral, borrow, vault_total_supply,
            veris_balance, veris_pct, net_rate, veris_aa_falconx, tranche_price,
            period_days, opening_value, interest, running_balance,
            tp_reengineered, collateral_usd, net, veris_share
        ))

        if running_balance is not None:
            prev_running_balance = running_balance
        prev_ts = ts

    # Insert into SQLite
    conn.executemany("""
        INSERT OR REPLACE INTO gauntlet_levered (
            timestamp_utc, block, collateral, borrow, vault_total_supply,
            veris_balance, veris_pct, net_rate, veris_aa_falconx, tranche_price,
            period_days, opening_value, interest, running_balance,
            tp_reengineered, collateral_usd, net, veris_share
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, processed)
    conn.commit()

    # Report
    last = processed[-1]
    print(f"  Inserted {len(processed)} rows into gauntlet_levered")
    print(f"  Last timestamp: {last[0]}")
    print(f"  Last running_balance: {last[13]:,.2f}" if last[13] else "  Last running_balance: None")
    print(f"  Last veris_share: {last[17]:,.2f}" if last[17] else "  Last veris_share: None")


def import_direct(conn, wb):
    """Import Direct Accrual sheet into direct_accrual table."""
    ws = wb['Direct Accrual']

    raw_rows = []
    for row in ws.iter_rows(min_row=2):
        ts = row[0].value
        if ts is None:
            continue
        # Skip section header rows (non-datetime values like "Period 3: ...")
        if isinstance(ts, str) and not ts[0].isdigit():
            continue
        raw_rows.append([cell.value for cell in row])

    if not raw_rows:
        print("Direct Accrual: no data rows found")
        return

    print(f"Direct Accrual: {len(raw_rows)} data rows read from xlsx")

    processed = []
    prev_running_balance = None
    prev_ts = None

    for row in raw_rows:
        ts = row[0]
        # Ensure ts is a datetime for arithmetic
        if isinstance(ts, str):
            ts = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
        token_balance = row[1] if row[1] is not None else VERIS_AA_TOKENS_DIRECT
        tranche_price = row[2]
        opening_value = row[3]
        net_rate = row[4]
        period_days = row[5]
        interest = row[6]
        running_balance = row[7]
        tp_reengineered = row[8] if len(row) > 8 else None

        # Compute period_days if missing
        if period_days is None and prev_ts is not None:
            delta = (ts - prev_ts).total_seconds()
            period_days = delta / 86400.0

        # Compute opening_value if missing
        if opening_value is None and prev_running_balance is not None:
            opening_value = prev_running_balance
        elif opening_value is None and prev_running_balance is None:
            # First row — use DIRECT_OPENING_VALUE
            opening_value = DIRECT_OPENING_VALUE

        # Compute interest if missing
        if interest is None and opening_value is not None and net_rate is not None and period_days is not None:
            interest = opening_value * net_rate * period_days / 365.0

        # Compute running_balance if missing
        if running_balance is None:
            if opening_value is not None and interest is not None:
                running_balance = opening_value + interest
            elif opening_value is not None:
                running_balance = opening_value

        # Compute tp_reengineered if missing
        if tp_reengineered is None and running_balance is not None and token_balance and token_balance > 0:
            tp_reengineered = running_balance / token_balance

        processed.append((
            _ts_to_iso(ts), token_balance, tranche_price,
            opening_value, net_rate, period_days, interest,
            running_balance, tp_reengineered
        ))

        if running_balance is not None:
            prev_running_balance = running_balance
        prev_ts = ts

    conn.executemany("""
        INSERT OR REPLACE INTO direct_accrual (
            timestamp_utc, token_balance, tranche_price,
            opening_value, net_rate, period_days, interest,
            running_balance, tp_reengineered
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, processed)
    conn.commit()

    last = processed[-1]
    print(f"  Inserted {len(processed)} rows into direct_accrual")
    print(f"  Last timestamp: {last[0]}")
    print(f"  Last running_balance: {last[7]:,.2f}" if last[7] else "  Last running_balance: None")
    print(f"  Last tp_reengineered: {last[8]:.6f}" if last[8] else "  Last tp_reengineered: None")


def export_to_csv(table_name, output_path=None):
    """Export a FalconX SQLite table to CSV."""
    if output_path is None:
        output_path = os.path.join(PROJECT_ROOT, 'outputs', f'falconx_{table_name}.csv')

    if not os.path.exists(DB_PATH):
        print(f"Database not found: {DB_PATH}")
        return

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.execute(f"SELECT * FROM {table_name} ORDER BY timestamp_utc")
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    conn.close()

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        writer.writerows(rows)

    print(f"Exported {len(rows)} rows to {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Import FalconX xlsx data to SQLite")
    parser.add_argument("--export", metavar="TABLE", help="Export table to CSV instead of importing (gauntlet_levered or direct_accrual)")
    args = parser.parse_args()

    if args.export:
        export_to_csv(args.export)
        return

    if not os.path.exists(XLSX_PATH):
        print(f"xlsx not found: {XLSX_PATH}")
        sys.exit(1)

    # Ensure data/ directory exists
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

    print(f"Reading {XLSX_PATH} (data_only=True)...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    print(f"Sheets: {wb.sheetnames}")

    conn = sqlite3.connect(DB_PATH)
    create_tables(conn)

    import_gauntlet(conn, wb)
    print()
    import_direct(conn, wb)

    wb.close()
    conn.close()

    print(f"\nDatabase written to {DB_PATH}")
    db_size = os.path.getsize(DB_PATH)
    print(f"Database size: {db_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
