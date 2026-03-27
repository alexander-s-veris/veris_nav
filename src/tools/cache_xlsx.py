"""
Cache xlsx sheets as individual CSVs for fast repeated access.

Large xlsx files (6MB+) take 30-60s to parse with openpyxl.
This script reads the file once and writes each sheet as a CSV,
plus a metadata JSON with sheet names, dimensions, and cache time.

Usage:
    python src/cache_xlsx.py <path_to_xlsx> [--output-dir cache/]

If --output-dir is not specified, defaults to cache/ in the project root.
Re-running overwrites existing cache for that file.
"""
import os
import sys
import csv
import json
import hashlib
from datetime import datetime, timezone

import openpyxl


def get_file_hash(filepath):
    """MD5 hash of the file to detect changes."""
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def sanitize_sheet_name(name):
    """Convert sheet name to a safe filename."""
    safe = name.replace("/", "_").replace("\\", "_").replace(":", "_")
    safe = safe.replace("*", "_").replace("?", "_").replace('"', "_")
    safe = safe.replace("<", "_").replace(">", "_").replace("|", "_")
    safe = safe.replace(" ", "_")
    # Collapse multiple underscores
    while "__" in safe:
        safe = safe.replace("__", "_")
    return safe.strip("_")


def cache_xlsx(xlsx_path, output_dir):
    """Read xlsx and write each sheet as CSV + metadata JSON."""
    xlsx_path = os.path.abspath(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    file_hash = get_file_hash(xlsx_path)
    file_size = os.path.getsize(xlsx_path)
    basename = os.path.basename(xlsx_path)

    print(f"Loading {basename} ({file_size / 1024 / 1024:.1f} MB)...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    metadata = {
        "source_file": basename,
        "source_path": xlsx_path,
        "source_hash_md5": file_hash,
        "source_size_bytes": file_size,
        "cached_at_utc": datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M:%S"),
        "sheet_count": len(wb.sheetnames),
        "sheets": {},
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sanitize_sheet_name(sheet_name)
        csv_filename = f"{safe_name}.csv"
        csv_path = os.path.join(output_dir, csv_filename)

        row_count = 0
        max_cols = 0

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                str_row = [str(v) if v is not None else "" for v in row]
                writer.writerow(str_row)
                row_count += 1
                max_cols = max(max_cols, len(str_row))

        metadata["sheets"][sheet_name] = {
            "csv_file": csv_filename,
            "rows": row_count,
            "cols": max_cols,
        }

        print(f"  {sheet_name} -> {csv_filename} ({row_count} rows, {max_cols} cols)")

    wb.close()

    # Write metadata
    meta_path = os.path.join(output_dir, "_metadata.json")
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)

    print(f"\nCached {len(wb.sheetnames)} sheets to {output_dir}/")
    print(f"Metadata: {meta_path}")
    return metadata


def is_cache_fresh(xlsx_path, output_dir):
    """Check if cache exists and matches the current file hash."""
    meta_path = os.path.join(output_dir, "_metadata.json")
    if not os.path.exists(meta_path):
        return False
    with open(meta_path) as f:
        meta = json.load(f)
    current_hash = get_file_hash(xlsx_path)
    return meta.get("source_hash_md5") == current_hash


def main():
    if len(sys.argv) < 2:
        print("Usage: python src/cache_xlsx.py <path_to_xlsx> [--output-dir cache/]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    cache_root = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "cache")

    for i, arg in enumerate(sys.argv):
        if arg == "--output-dir" and i + 1 < len(sys.argv):
            cache_root = sys.argv[i + 1]

    if not os.path.exists(xlsx_path):
        print(f"Error: file not found: {xlsx_path}")
        sys.exit(1)

    # Create subfolder named after the xlsx file (without extension)
    xlsx_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    output_dir = os.path.join(cache_root, xlsx_name)

    if is_cache_fresh(xlsx_path, output_dir):
        print(f"Cache is up to date for {os.path.basename(xlsx_path)}")
        print(f"Cached at: {output_dir}/")
        print(f"Delete {output_dir}/_metadata.json to force re-cache.")
        return

    cache_xlsx(xlsx_path, output_dir)


if __name__ == "__main__":
    main()
